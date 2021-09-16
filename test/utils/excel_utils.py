import openpyxl
from openpyxl.styles import Alignment


def get_row_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_row_count_specific_column(file, sheet_name, col):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return len(sheet[col])


def get_col_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_column


def read_data(file, sheet_name, row_no, col_no):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row_no, col_no).value


def read_data_without_open_file(sheet, row_num, col_num):
    return sheet.cell(row_num, col_num).value


def write_data(file, sheet_name, row_num, col_num, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(row_num, col_num).value = data
    workbook.save(file)


def read_single_row(file, sheet_name, row_num, col_num):
    single_row = []
    for i in range(1, col_num + 1):
        data = read_data(file, sheet_name, row_num, i)
        single_row.append(data)
    return single_row


def read_single_col(file, sheet_name, starting_row, max_row_num, col_num):
    single_col = []
    for i in range(starting_row, max_row_num + 1):
        data = read_data(file, sheet_name, i, col_num)
        single_col.append(data)
    return single_col


def read_single_col_open_file_once(file, sheet_name, starting_row, max_row_num, col_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    single_col = []
    for i in range(starting_row, max_row_num + 1):
        data = read_data_without_open_file(sheet, i, col_num)
        single_col.append(data)
    return single_col


def write_list_of_list(file, sheet_name, data):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheet_name]
    row = 1
    for element in data:
        col = 1
        for d in element:
            if d:
                sheet.cell(row, col).value = d
            col += 1
        row += 1
    wb.save(file)


def write_single_row(file, sheet_name, row_num, col_num, starting_col, data):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheet_name]
    for i in range(1, col_num + 1):
        sheet.cell(row_num, i + starting_col).value = data[i - 1]
        # print(data[i - 1])
    wb.save(file)


def write_single_col(file, sheet_name, max_row_num, col_num, starting_row, data):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheet_name]
    for i in range(1, max_row_num + 1):
        sheet.cell(i + starting_row, col_num).value = str(data[i - 1])
        # print(data[i - 1])
    wb.save(file)


def merge_cell(file, sheet_name, starting_merge_cell, ending_merge_cell, row_num, col_num, ):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheet_name]
    sheet.merge_cells(f'{starting_merge_cell}:{ending_merge_cell}')
    cell = sheet.cell(row_num, col_num)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    wb.save(file)


def write_in_merged_cell(file, sheet_name, starting_merge_cell, ending_merge_cell, row_num, col_num, starting_col,
                         data):
    merge_cell(file, sheet_name, starting_merge_cell, ending_merge_cell, row_num, starting_col + 1)
    write_single_row(file, sheet_name, row_num, col_num, starting_col, data)


def remove_items(list1, item):
    res = [i for i in list1 if i != item]
    return res


# readonly

def get_row_count_read_only(file, sheet_name):
    workbook = openpyxl.load_workbook(filename=file, read_only=True)
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value)
    # max_row = sheet.max_row
    # print(max_row)
    workbook.close()
    # return max_row


def get_col_count_read_only(file, sheet_name):
    workbook = openpyxl.load_workbook(file, read_only=True)
    sheet = workbook[sheet_name]
    max_column = sheet.max_column
    workbook.close()
    return max_column


def read_single_col_read_only(file, sheet_name, starting_row, max_row_num, col_num):
    single_col = []
    for i in range(starting_row, max_row_num + 1):
        data = read_data(file, sheet_name, i, col_num)
        single_col.append(data)
    return single_col
